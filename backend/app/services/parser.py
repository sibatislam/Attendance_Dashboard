from __future__ import annotations

from typing import List, Dict, Any, Tuple
import io
import csv
from openpyxl import load_workbook

try:
    import xlrd  # for .xls legacy support
except Exception:  # pragma: no cover
    xlrd = None


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _read_csv(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows_iter = iter(reader)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []
    header_order = [str(h) for h in header]
    rows: List[Dict[str, Any]] = []
    for r in rows_iter:
        record = {}
        for idx, col in enumerate(header_order):
            record[col] = _stringify(r[idx]) if idx < len(r) else ""
        rows.append(record)
    return header_order, rows


def _read_xlsx(file_bytes: bytes, data_only: bool = False) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=data_only, read_only=True)
    ws = wb.active
    first = True
    header_order: List[str] = []
    rows: List[Dict[str, Any]] = []
    for row_cells in ws.iter_rows(values_only=False):
        values = [_stringify(c.value) for c in row_cells]
        if first:
            header_order = [str(h) for h in values]
            first = False
            continue
        if not header_order:
            continue
        record = {}
        for idx, col in enumerate(header_order):
            record[col] = values[idx] if idx < len(values) else ""
        rows.append(record)
    return header_order, rows


def _read_xls(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    if xlrd is None:
        raise ValueError(".xls support requires xlrd; please install xlrd==1.2.0")
    book = xlrd.open_workbook(file_contents=file_bytes)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return [], []
    header_order = [str(sheet.cell_value(0, c)) for c in range(sheet.ncols)]
    rows: List[Dict[str, Any]] = []
    for r in range(1, sheet.nrows):
        record = {}
        for c in range(sheet.ncols):
            record[header_order[c]] = _stringify(sheet.cell_value(r, c))
        rows.append(record)
    return header_order, rows


def read_file_preserve_text(filename: str, file_bytes: bytes, data_only: bool = False) -> Tuple[List[str], List[Dict[str, Any]]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return _read_csv(file_bytes)
    if lower_name.endswith(".xlsx"):
        return _read_xlsx(file_bytes, data_only=data_only)
    if lower_name.endswith(".xls"):
        return _read_xls(file_bytes)
    raise ValueError("Unsupported file type. Please upload CSV or Excel files.")


# Standard column names for Teams user list merge
_TEAMS_USER_LIST_STANDARD = ["Name", "Email", "Designation", "Department", "Function"]
# Map known header variants (case-insensitive) to standard name by position
_TEAMS_HEADER_ALIASES = {
    "name": "Name", "col1": "Name",
    "email": "Email", "col2": "Email",
    "designation": "Designation", "col3": "Designation",
    "department": "Department", "col4": "Department",
    "function": "Function", "col5": "Function",
}


def _normalize_teams_row(raw: Dict[str, Any], sheet_label: str) -> Dict[str, Any]:
    """Convert a row with arbitrary headers to standard keys. Drops serial/other columns."""
    out = {"Sheet": sheet_label}
    used = set()
    for key, value in raw.items():
        if key == "Sheet":
            continue
        key_lower = str(key).strip().lower()
        if not key_lower:
            continue
        # Match standard names or COL1..COL5 / Name, Email, etc.
        for alias, standard in _TEAMS_HEADER_ALIASES.items():
            if key_lower == alias or key_lower.replace(" ", "") == alias:
                if standard not in used:
                    out[standard] = _stringify(value)
                    used.add(standard)
                break
        else:
            if key_lower.startswith("col") and len(key_lower) <= 5:
                idx = key_lower.replace("col", "").strip()
                if idx == "1":
                    out["Name"] = _stringify(value)
                elif idx == "2":
                    out["Email"] = _stringify(value)
                elif idx == "3":
                    out["Designation"] = _stringify(value)
                elif idx == "4":
                    out["Department"] = _stringify(value)
                elif idx == "5":
                    out["Function"] = _stringify(value)
    for col in _TEAMS_USER_LIST_STANDARD:
        if col not in out:
            out[col] = ""
    return out


def _is_teams_header_row(normalized: Dict[str, Any]) -> bool:
    """True if this row is the sheet header row (Name, Email, Designation, etc. as values)."""
    vals = [str(normalized.get(c) or "").strip().lower() for c in _TEAMS_USER_LIST_STANDARD]
    standard_lower = [c.lower() for c in _TEAMS_USER_LIST_STANDARD]
    return vals == standard_lower


# Only count rows that have MICROSOFT TEAMS ESSENTIALS in Assigned Products (if column exists)
_TEAMS_ESSENTIALS = "microsoft teams essentials"


def _row_has_teams_essentials(row_dict: Dict[str, Any]) -> bool:
    """True if row has an Assigned Products column and it contains MICROSOFT TEAMS ESSENTIALS."""
    for key, value in row_dict.items():
        key_lower = str(key).strip().lower()
        if "assigned" in key_lower and "product" in key_lower:
            val = _stringify(value).strip().lower()
            if _TEAMS_ESSENTIALS in val:
                return True
            return False
    return True  # No Assigned Products column: include row (backward compatibility)


def _row_is_licensed_yes(row_dict: Dict[str, Any]) -> bool:
    """True if row has no 'Is Licensed' column or its value is 'Yes' (case-insensitive)."""
    for key, value in row_dict.items():
        key_lower = str(key).strip().lower()
        if key_lower.replace(" ", "") == "islicensed":
            val = _stringify(value).strip().lower()
            return val == "yes"
    return True  # No Is Licensed column: include row (backward compatibility)


def read_teams_user_list_sheets(file_bytes: bytes) -> Dict[str, Any]:
    """
    Read Excel file and count data rows + return row data from sheets 'Teams' and 'CBL_Teams'.
    Rows are normalized to standard columns (Name, Email, Designation, Department, Function, Sheet)
    and given a continuous S.No. Blank rows are skipped.
    Returns dict:
      by_sheet: { 'Teams': count, 'CBL_Teams': count }
      total_assigned: int
      rows: list of dicts with S.No, Sheet, Name, Email, Designation, Department, Function
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}
    by_sheet = {}
    all_rows: List[Dict[str, Any]] = []
    for want in ("teams", "cbl_teams"):
        sheet_label = "CBL_Teams" if want == "cbl_teams" else "Teams"
        if want not in sheet_names_lower:
            by_sheet[sheet_label] = 0
            continue
        name = sheet_names_lower[want]
        ws = wb[name]
        header_order: List[str] = []
        count = 0
        for i, row_cells in enumerate(ws.iter_rows(values_only=True)):
            values = [c for c in row_cells]
            if i == 0:
                header_order = [str(h or "").strip() or f"Col{j}" for j, h in enumerate(values)]
                continue
            if not header_order:
                continue
            row_dict = {}
            for idx, col in enumerate(header_order):
                row_dict[col] = _stringify(values[idx]) if idx < len(values) else ""
            if not any(v for v in row_dict.values() if str(v).strip()):
                continue
            if not _row_has_teams_essentials(row_dict):
                continue
            if not _row_is_licensed_yes(row_dict):
                continue
            normalized = _normalize_teams_row(row_dict, sheet_label)
            if _is_teams_header_row(normalized):
                continue
            if any(normalized.get(c) for c in _TEAMS_USER_LIST_STANDARD if str(normalized.get(c) or "").strip()):
                all_rows.append(normalized)
                count += 1
        by_sheet[sheet_label] = count
    wb.close()

    for idx, row in enumerate(all_rows, start=1):
        row["S.No"] = idx

    return {
        "by_sheet": by_sheet,
        "total_assigned": len(all_rows),
        "rows": all_rows,
    }


