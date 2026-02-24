"""User management endpoints (admin only)."""
import io
from typing import List

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from ..db import get_db
from ..models import User, Role
from ..schemas import UserResponse, UserCreate, UserUpdate
from ..auth import get_current_admin_user, get_current_user, get_password_hash
from ..utils import datetime_to_iso_utc
from .roles import get_permissions_for_role, _ensure_default_roles
from ..services.employee_hierarchy import build_hierarchy_map, scope_for_user, get_scope_options_for_user, scope_to_persist_for_user


class BulkDeleteRequest(BaseModel):
    user_ids: List[int]

router = APIRouter()

BULK_DEFAULT_PASSWORD = "123456"
TEMPLATE_HEADERS = [
    "Employee Name",
    "Designation",
    "Function",
    "Email (Official)",
    "Username",
    "Role",
    "Password",
]
EMAIL_COL_VARIANTS = ["Email (Official)", "Email (Offical)"]


@router.get("/")
def list_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """List all users (admin only). Permissions resolved from Role by user.role."""
    users = db.query(User).all()
    out = []
    for u in users:
        role_name = u.role or "user"
        perms = get_permissions_for_role(db, role_name)
        if not isinstance(perms, dict):
            perms = {}
        out.append({
            "id": u.id,
            "email": u.email or "",
            "username": u.username or "",
            "full_name": u.full_name,
            "phone": u.phone,
            "department": u.department,
            "position": u.position,
            "role": role_name,
            "is_active": bool(u.is_active),
            "permissions": perms,
            "employee_email": u.employee_email or None,
            "data_scope_level": u.data_scope_level or None,
            "allowed_functions": u.allowed_functions if isinstance(u.allowed_functions, list) else ([] if u.allowed_functions is None else list(u.allowed_functions)),
            "allowed_departments": u.allowed_departments if isinstance(u.allowed_departments, list) else ([] if u.allowed_departments is None else list(u.allowed_departments)),
            "allowed_companies": u.allowed_companies if isinstance(u.allowed_companies, list) else ([] if u.allowed_companies is None else list(u.allowed_companies)),
            "last_login": datetime_to_iso_utc(u.last_login),
            "created_at": datetime_to_iso_utc(u.created_at),
            "updated_at": datetime_to_iso_utc(u.updated_at),
        })
    return out


@router.get("/me/scope")
def get_my_scope(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Return data scope and tab visibility for current user.
    - Admin -> { "all": true, "visible_tabs": ["function", "company", "location", "department"] }.
    - If user has allowed_functions / allowed_departments / allowed_companies set (non-empty), use those.
    - Else use hierarchy (N / N-1 / N-2) or own from employee.
    - visible_tabs: which Weekly Dashboard tabs the user can see (from role permissions tab_function, tab_company, etc.).
    """
    perms = get_permissions_for_role(db, current_user.role or "user")
    att = (perms or {}).get("attendance_dashboard") or {}
    features = att.get("features") or []
    
    # Weekly Dashboard tabs: use ONLY menu-specific IDs so role checkboxes control visibility.
    # (Legacy tab_* IDs are not used here; otherwise roles created with defaults would always show all tabs.)
    visible_tabs = []
    weekly_tab_checks = [
        ("weekly_dashboard_tab_function", "function"),
        ("weekly_dashboard_tab_company", "company"),
        ("weekly_dashboard_tab_location", "location"),
        ("weekly_dashboard_tab_department", "department"),
    ]
    for feat_id, tab_name in weekly_tab_checks:
        if feat_id in features:
            visible_tabs.append(tab_name)
    if not visible_tabs:
        visible_tabs = ["function", "company", "location", "department"]

    # Dashboard tabs: use ONLY dashboard_tab_* IDs (no legacy) so role checkboxes control visibility.
    visible_tabs_dashboard = []
    dashboard_tab_checks = [
        ("dashboard_tab_function", "function"),
        ("dashboard_tab_company", "company"),
        ("dashboard_tab_location", "location"),
    ]
    for feat_id, tab_name in dashboard_tab_checks:
        if feat_id in features:
            visible_tabs_dashboard.append(tab_name)
    if not visible_tabs_dashboard:
        visible_tabs_dashboard = ["function", "company", "location"]

    # User Wise page tabs: use ONLY user_wise_* IDs (no legacy) so role checkboxes control visibility.
    visible_tabs_user_wise = []
    user_wise_tab_checks = [
        ("user_wise_on_time", "on_time"),
        ("user_wise_work_hour", "work_hour"),
        ("user_wise_work_hour_lost", "work_hour_lost"),
        ("user_wise_work_hour_lost_cost", "work_hour_lost_cost"),
        ("user_wise_leave_analysis", "leave_analysis"),
        ("user_wise_od_analysis", "od_analysis"),
    ]
    for feat_id, tab_name in user_wise_tab_checks:
        if feat_id in features:
            visible_tabs_user_wise.append(tab_name)
    if not visible_tabs_user_wise:
        visible_tabs_user_wise = ["on_time", "work_hour", "work_hour_lost", "work_hour_lost_cost", "leave_analysis", "od_analysis"]

    filter_options = get_scope_options_for_user(db, current_user)
    if current_user.role == "admin":
        return {
            "all": True,
            "allowed_functions": None,
            "allowed_departments": None,
            "allowed_companies": None,
            "data_scope_level": None,
            "visible_tabs": visible_tabs,
            "visible_tabs_dashboard": visible_tabs_dashboard,
            "visible_tabs_user_wise": visible_tabs_user_wise,
            "filter_options": filter_options,
        }
    # User-level allowed lists: if set (non-empty), use them; else derive from hierarchy / own
    af = getattr(current_user, "allowed_functions", None)
    ad = getattr(current_user, "allowed_departments", None)
    ac = getattr(current_user, "allowed_companies", None)
    if isinstance(af, list) and len(af) > 0:
        return {
            "all": False,
            "allowed_functions": af,
            "allowed_departments": ad if isinstance(ad, list) else None,
            "allowed_companies": ac if isinstance(ac, list) else None,
            "data_scope_level": current_user.data_scope_level,
            "visible_tabs": visible_tabs,
            "visible_tabs_dashboard": visible_tabs_dashboard,
            "visible_tabs_user_wise": visible_tabs_user_wise,
            "filter_options": filter_options,
        }
    if isinstance(ad, list) and len(ad) > 0:
        return {
            "all": False,
            "allowed_functions": af if isinstance(af, list) else None,
            "allowed_departments": ad,
            "allowed_companies": ac if isinstance(ac, list) else None,
            "data_scope_level": current_user.data_scope_level,
            "visible_tabs": visible_tabs,
            "visible_tabs_dashboard": visible_tabs_dashboard,
            "visible_tabs_user_wise": visible_tabs_user_wise,
            "filter_options": filter_options,
        }
    if isinstance(ac, list) and len(ac) > 0:
        return {
            "all": False,
            "allowed_functions": af if isinstance(af, list) else None,
            "allowed_departments": ad if isinstance(ad, list) else None,
            "allowed_companies": ac,
            "data_scope_level": current_user.data_scope_level,
            "visible_tabs": visible_tabs,
            "visible_tabs_dashboard": visible_tabs_dashboard,
            "visible_tabs_user_wise": visible_tabs_user_wise,
            "filter_options": filter_options,
        }
    hierarchy_map = build_hierarchy_map(db, None)
    out = scope_for_user(
        db,
        current_user.employee_email,
        current_user.data_scope_level,
        hierarchy_map=hierarchy_map,
    )
    out["visible_tabs"] = visible_tabs
    out["visible_tabs_dashboard"] = visible_tabs_dashboard
    out["visible_tabs_user_wise"] = visible_tabs_user_wise
    if "allowed_companies" not in out:
        out["allowed_companies"] = None
    out["filter_options"] = filter_options
    return out


@router.post("/", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    user_data: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Create a new user (admin only)."""
    # Check if email exists
    existing_email = db.query(User).filter(User.email == user_data.email).first()
    if existing_email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )
    
    # Check if username exists
    existing_username = db.query(User).filter(User.username == user_data.username).first()
    if existing_username:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username already taken"
        )
    
    # Validate role exists
    _ensure_default_roles(db)
    role_name = (user_data.role or "user").strip()
    if not db.query(Role).filter(Role.name == role_name).first():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Role '{role_name}' not found. Create it in Role Management first."
        )
    hashed_password = get_password_hash(user_data.password)
    new_user = User(
        email=user_data.email,
        username=user_data.username,
        full_name=user_data.full_name,
        phone=user_data.phone,
        department=user_data.department,
        position=user_data.position,
        hashed_password=hashed_password,
        role=role_name,
        permissions={},
        employee_email=getattr(user_data, "employee_email", None) or None,
        data_scope_level=getattr(user_data, "data_scope_level", None) or None,
        allowed_functions=getattr(user_data, "allowed_functions", None) or [],
        allowed_departments=getattr(user_data, "allowed_departments", None) or [],
        allowed_companies=getattr(user_data, "allowed_companies", None) or [],
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return new_user


def _username_from_email(email: str) -> str:
    """Extract username as part before @ (e.g. user@cg-bd.com -> user)."""
    if not email or "@" not in email:
        return ""
    return email.strip().split("@")[0].strip()


def _find_header_index(headers: List[str], names: List[str]) -> int:
    """Return 0-based column index for first matching header (case-insensitive, stripped)."""
    normalized = [h.strip().lower() if h else "" for h in headers]
    keys = {n.strip().lower() for n in names}
    for i, h in enumerate(normalized):
        if h in keys:
            return i
    return -1


@router.get("/bulk-upload/template")
def download_bulk_upload_template(
    current_user: User = Depends(get_current_admin_user),
):
    """Download Excel template for bulk user upload (admin only)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(TEMPLATE_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=bulk_users_template.xlsx"},
    )


@router.post("/bulk-upload")
def bulk_upload_users(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user),
) -> dict:
    """Bulk create users from Excel. Default password: 123456. Columns: Employee Name, Designation, Function, Email (Official)/Email (Offical), Username (optional, derived from email if missing)."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be .xlsx",
        )

    content = file.file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty file")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid Excel file: {str(e)}",
        )

    if len(rows) < 2:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must have a header row and at least one data row",
        )

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx_name = _find_header_index(headers, ["Employee Name"])
    idx_desig = _find_header_index(headers, ["Designation"])
    idx_func = _find_header_index(headers, ["Function"])
    idx_email = _find_header_index(headers, EMAIL_COL_VARIANTS)
    idx_username = _find_header_index(headers, ["Username"])
    idx_role = _find_header_index(headers, ["Role"])
    idx_password = _find_header_index(headers, ["Password"])

    if idx_email < 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Missing 'Email (Official)' or 'Email (Offical)' column",
        )

    created: List[dict] = []
    skipped: List[dict] = []
    errors: List[dict] = []

    def _cell(row: tuple, i: int) -> str:
        if i < 0 or i >= len(row):
            return ""
        v = row[i]
        return (str(v).strip() if v is not None else "") or ""

    for row_idx, row in enumerate(rows[1:], start=2):
        raw = tuple(row) if isinstance(row, (list, tuple)) else (row,)
        
        email = _cell(raw, idx_email)
        email_raw = email  # Keep original for error messages
        
        # Skip rows where email is empty (likely empty/invalid rows)
        if not email or not email.strip():
            # Only report as error if at least one other field has data (partial row)
            has_other_data = (
                (idx_name >= 0 and _cell(raw, idx_name)) or
                (idx_desig >= 0 and _cell(raw, idx_desig)) or
                (idx_func >= 0 and _cell(raw, idx_func))
            )
            if has_other_data:
                errors.append({"row": row_idx, "reason": "Missing or empty email", "email_value": "(empty)"})
            continue
        
        if "@" not in email:
            errors.append({"row": row_idx, "reason": f"Invalid email format (missing @): {email_raw}", "email_value": email_raw})
            continue

        email = email.strip().lower()
        
        # Validate email format more strictly
        if not email or len(email.split("@")) != 2:
            errors.append({"row": row_idx, "reason": f"Invalid email format: {email_raw}", "email_value": email_raw})
            continue
        
        username = _cell(raw, idx_username) if idx_username >= 0 else ""
        if not username:
            username = _username_from_email(email)
        else:
            username = username.strip()
        if not username:
            errors.append({"row": row_idx, "reason": f"Could not derive username from email: {email_raw}", "email_value": email_raw})
            continue

        full_name = _cell(raw, idx_name) if idx_name >= 0 else ""
        position = _cell(raw, idx_desig) if idx_desig >= 0 else ""
        department = _cell(raw, idx_func) if idx_func >= 0 else ""
        role_raw = _cell(raw, idx_role) if idx_role >= 0 else ""
        _ensure_default_roles(db)
        role_name_stripped = role_raw.strip()
        if role_name_stripped.lower() == "admin":
            role = "admin"
        elif role_name_stripped and db.query(Role).filter(Role.name == role_name_stripped).first():
            role = role_name_stripped
        else:
            role = "user"
        password_raw = _cell(raw, idx_password) if idx_password >= 0 else ""
        password = password_raw if password_raw else BULK_DEFAULT_PASSWORD
        hashed = get_password_hash(password)

        existing_email = db.query(User).filter(User.email == email).first()
        if existing_email:
            skipped.append({"row": row_idx, "email": email, "reason": "Email already registered"})
            continue
        existing_username = db.query(User).filter(User.username == username).first()
        if existing_username:
            skipped.append({"row": row_idx, "username": username, "reason": "Username already taken"})
            continue

        try:
            u = User(
                email=email,
                username=username,
                full_name=full_name or None,
                phone=None,
                department=department or None,
                position=position or None,
                hashed_password=hashed,
                role=role,
                permissions={},
            )
            db.add(u)
            db.commit()
            db.refresh(u)
            created.append({"row": row_idx, "email": email, "username": username})
        except Exception as e:
            db.rollback()
            errors.append({"row": row_idx, "email": email, "reason": str(e)})

    return {
        "created": len(created),
        "skipped": len(skipped),
        "errors": len(errors),
        "details": {"created": created, "skipped": skipped, "errors": errors},
    }


@router.post("/bulk-delete")
def bulk_delete_users(
    body: BulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user),
):
    """Delete multiple users (admin only). Skips self, admins, and non-existent IDs."""
    deleted = 0
    skipped = []
    for uid in body.user_ids:
        if uid == current_user.id:
            skipped.append({"id": uid, "reason": "Cannot delete your own account"})
            continue
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            skipped.append({"id": uid, "reason": "User not found"})
            continue
        if user.role == "admin":
            skipped.append({"id": uid, "reason": "Cannot delete admin users"})
            continue
        db.delete(user)
        deleted += 1
    db.commit()
    return {"deleted": deleted, "skipped": skipped}


@router.get("/{user_id}", response_model=UserResponse)
def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Get a specific user (admin only)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    return user


@router.put("/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    user_data: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Update a user (admin only)."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Update fields if provided
    if user_data.email is not None:
        # Check if new email is already taken
        existing = db.query(User).filter(
            User.email == user_data.email,
            User.id != user_id
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Email already in use"
            )
        user.email = user_data.email
    
    if user_data.username is not None:
        # Check if new username is already taken
        existing = db.query(User).filter(
            User.username == user_data.username,
            User.id != user_id
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Username already in use"
            )
        user.username = user_data.username
    
    if user_data.full_name is not None:
        user.full_name = user_data.full_name
    
    if user_data.phone is not None:
        user.phone = user_data.phone
    
    if user_data.department is not None:
        user.department = user_data.department
    
    if user_data.position is not None:
        user.position = user_data.position
    
    if user_data.password is not None:
        user.hashed_password = get_password_hash(user_data.password)
    
    if user_data.role is not None and user_data.role.strip():
        _ensure_default_roles(db)
        rn = user_data.role.strip()
        role_exists = db.query(Role).filter(Role.name == rn).first() is not None
        if not role_exists:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Role '{rn}' not found. Create it in Role Management first."
            )
        user.role = rn
    
    if user_data.is_active is not None:
        user.is_active = user_data.is_active

    if hasattr(user_data, "employee_email") and user_data.employee_email is not None:
        user.employee_email = user_data.employee_email.strip() or None
    if hasattr(user_data, "employee_email") and user_data.employee_email is not None and user_data.employee_email.strip() == "":
        user.employee_email = None
    if hasattr(user_data, "data_scope_level"):
        user.data_scope_level = (user_data.data_scope_level or "").strip() or None
    if hasattr(user_data, "allowed_functions"):
        user.allowed_functions = user_data.allowed_functions if user_data.allowed_functions is not None else []
    if hasattr(user_data, "allowed_departments"):
        user.allowed_departments = user_data.allowed_departments if user_data.allowed_departments is not None else []
    if hasattr(user_data, "allowed_companies"):
        user.allowed_companies = user_data.allowed_companies if user_data.allowed_companies is not None else []

    db.commit()
    db.refresh(user)
    
    return user


def _is_n_level(level: str) -> bool:
    """Check if level is N, N-1, N-2, N-3, etc."""
    if not level:
        return False
    level = level.strip()
    if level == "N":
        return True
    import re
    return bool(re.match(r"^N-\d+$", level))


@router.get("/scope-from-hierarchy")
def get_scope_from_hierarchy(
    employee_email: str = Query(..., description="Employee email (Official) from Employee List"),
    data_scope_level: str = Query(..., description="N, N-1, N-2, etc."),
    employee_file_id: int | None = Query(None, description="Optional: use specific Employee List file; else latest"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user),
):
    """
    Return allowed_companies, allowed_functions, allowed_departments for the given employee and level,
    computed from the Employee List hierarchy. Use in Edit User to pre-fill Allowed Companies/Functions/Departments.
    """
    hierarchy = build_hierarchy_map(db, employee_file_id)
    scope = scope_to_persist_for_user(hierarchy, employee_email, data_scope_level)
    return {
        "allowed_companies": scope.get("allowed_companies") or [],
        "allowed_functions": scope.get("allowed_functions") or [],
        "allowed_departments": scope.get("allowed_departments") or [],
    }


@router.post("/sync-roles-from-hierarchy")
def sync_roles_from_hierarchy(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user),
):
    """
    For each user with Employee Email matching the Employee List hierarchy,
    set their role and data_scope_level to their hierarchy level (N, N-1, N-2, N-3, ...).
    Skips admin users. Creates missing N-* roles automatically.
    """
    from .roles import _ensure_default_roles, DEFAULT_N_PERMISSIONS
    _ensure_default_roles(db)
    hierarchy = build_hierarchy_map(db, None)
    if not hierarchy:
        return {"message": "No employee hierarchy found. Upload an Employee List first.", "updated": 0}
    # Collect all unique N-* levels in hierarchy
    all_levels = {(emp.get("level") or "").strip() for emp in hierarchy.values()}
    n_levels = {lvl for lvl in all_levels if _is_n_level(lvl)}
    # Ensure roles exist for all N-* levels
    for lvl in n_levels:
        existing = db.query(Role).filter(Role.name == lvl).first()
        if not existing:
            db.add(Role(name=lvl, permissions=DEFAULT_N_PERMISSIONS))
    db.commit()
    updated = 0
    users = db.query(User).filter(User.role != "admin").all()
    for user in users:
        emp_email = (user.employee_email or user.email or "").strip()
        if not emp_email:
            continue
        emp = hierarchy.get(emp_email.lower())
        if not emp:
            continue
        level = (emp.get("level") or "").strip()
        if not _is_n_level(level):
            continue
        if not (user.employee_email or "").strip():
            user.employee_email = user.email
        role_or_scope_changed = (
            user.role != level
            or (user.data_scope_level or "").strip() != level
        )
        if role_or_scope_changed:
            user.role = level
            user.data_scope_level = level
        # Assign allowed_companies, allowed_functions, allowed_departments from hierarchy (N-1, N-2, etc.)
        scope = scope_to_persist_for_user(hierarchy, emp_email, level)
        user.allowed_companies = scope.get("allowed_companies") or []
        user.allowed_functions = scope.get("allowed_functions") or []
        user.allowed_departments = scope.get("allowed_departments") or []
        updated += 1
    db.commit()
    return {"message": f"Synced roles and data scope from hierarchy. {updated} user(s) updated. Company, Function, and Department assigned for N-1, N-2, etc.", "updated": updated}


@router.delete("/{user_id}")
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin_user)
):
    """Delete a user (admin only)."""
    # Prevent admin from deleting themselves
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete your own account"
        )
    
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User not found"
        )
    
    # Prevent deletion of admin users (protect all admin accounts)
    if user.role == "admin":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete admin users"
        )
    
    db.delete(user)
    db.commit()
    
    return {"message": "User deleted successfully"}

